import os

try:
    from barcode import Code128
    from barcode.writer import ImageWriter
    BARCODE_DISPONIBLE = True
except ImportError:
    BARCODE_DISPONIBLE = False

try:
    import qrcode
    QR_DISPONIBLE = True
except ImportError:
    QR_DISPONIBLE = False

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ImagenExcel
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    IMAGEN_DISPONIBLE = True
except ImportError:
    IMAGEN_DISPONIBLE = False

CARPETA_SALIDA = "salida"

def _asegurar_carpeta():
    if not os.path.exists(CARPETA_SALIDA):
        os.makedirs(CARPETA_SALIDA)

def _cargar_fuente(tamano=18):
    if not IMAGEN_DISPONIBLE:
        return None
    try:
        return ImageFont.truetype("arial.ttf", tamano)
    except Exception:
        return ImageFont.load_default()

def generar_codigo_barras(codigo):
    _asegurar_carpeta()
    ruta_base = os.path.join(CARPETA_SALIDA, "barcode_" + codigo)
    if BARCODE_DISPONIBLE:
        barras = Code128(codigo, writer=ImageWriter())
        return barras.save(ruta_base)
    ruta_txt = ruta_base + ".txt"
    with open(ruta_txt, "w", encoding="utf-8") as archivo:
        archivo.write("CODIGO DE BARRAS (texto): " + codigo + "\n")
    return ruta_txt

def generar_qr(contenido, nombre):
    _asegurar_carpeta()
    ruta = os.path.join(CARPETA_SALIDA, nombre)
    if QR_DISPONIBLE:
        qr = qrcode.QRCode(version=2, box_size=10, border=4)
        qr.add_data(contenido)
        qr.make(fit=True)
        imagen = qr.make_image(fill_color="black", back_color="white")
        ruta_png = ruta + ".png"
        imagen.save(ruta_png)
        return ruta_png
    ruta_txt = ruta + ".txt"
    with open(ruta_txt, "w", encoding="utf-8") as archivo:
        archivo.write(contenido)
    return ruta_txt

def generar_rotulado_imagen(texto, nombre):
    _asegurar_carpeta()
    ruta = os.path.join(CARPETA_SALIDA, nombre)
    if IMAGEN_DISPONIBLE:
        imagen = PILImage.new("RGB", (620, 460), color="white")
        dibujo = ImageDraw.Draw(imagen)
        dibujo.multiline_text((20, 20), texto, fill="black",
                              font=_cargar_fuente(), spacing=6)
        ruta_png = ruta + ".png"
        imagen.save(ruta_png)
        return ruta_png
    ruta_txt = ruta + ".txt"
    with open(ruta_txt, "w", encoding="utf-8") as archivo:
        archivo.write(texto)
    return ruta_txt

def _ajustar_ancho_columnas(hoja, encabezados):
    from openpyxl.utils import get_column_letter
    for indice, encabezado in enumerate(encabezados, start=1):
        letra = get_column_letter(indice)
        ancho = max(12, min(38, len(str(encabezado)) + 6))
        hoja.column_dimensions[letra].width = ancho

def exportar_excel(nombre_archivo, encabezados, filas, imagenes=None):
    _asegurar_carpeta()
    ruta = os.path.join(CARPETA_SALIDA, nombre_archivo)
    if EXCEL_DISPONIBLE:
        from openpyxl.utils import get_column_letter

        TAMANO_IMAGEN_PX = 100
        ANCHO_COLUMNA_IMAGEN = 15
        ALTO_FILA_IMAGEN = 78

        libro = Workbook()
        hoja = libro.active
        hoja.title = "Reporte"
        hoja.append(encabezados)
        for fila in filas:
            hoja.append(fila)
        _ajustar_ancho_columnas(hoja, encabezados)
        if imagenes:
            for ruta_img, fila_num, columna_num in imagenes:
                if ruta_img and os.path.exists(ruta_img) and ruta_img.lower().endswith(".png"):
                    letra_columna = get_column_letter(columna_num)
                    img = ImagenExcel(ruta_img)
                    img.width = TAMANO_IMAGEN_PX
                    img.height = TAMANO_IMAGEN_PX
                    hoja.column_dimensions[letra_columna].width = ANCHO_COLUMNA_IMAGEN
                    hoja.row_dimensions[fila_num].height = ALTO_FILA_IMAGEN
                    hoja.add_image(img, letra_columna + str(fila_num))
        ruta_xlsx = ruta + ".xlsx"
        libro.save(ruta_xlsx)
        return ruta_xlsx
    ruta_csv = ruta + ".csv"
    with open(ruta_csv, "w", encoding="utf-8") as archivo:
        archivo.write(";".join(str(c) for c in encabezados) + "\n")
        for fila in filas:
            archivo.write(";".join(str(c) for c in fila) + "\n")
    return ruta_csv

def librerias_faltantes():
    faltantes = []
    if not BARCODE_DISPONIBLE:
        faltantes.append("python-barcode")
    if not QR_DISPONIBLE:
        faltantes.append("qrcode")
    if not EXCEL_DISPONIBLE:
        faltantes.append("openpyxl")
    if not IMAGEN_DISPONIBLE:
        faltantes.append("pillow")
    return ", ".join(faltantes)
